import threading
from mttkinter import mtTkinter as mtk
from tkinter import ttk, Scrollbar, filedialog
from tkinter.messagebox import showinfo, showerror
from openpyxl import Workbook, load_workbook
import asyncio
import aiohttp
import os
import re
import random

class Application(object):

    def __init__(self):
        self.__createGUI()
        self.treeIndex = 1
        self.totals = 0
        self.totalData = [["用户ID", "OLT名称", "PON端口", "用户状态", "时间"]]
        self.excelData = []

    def __createGUI(self):
        self.root = mtk.Tk()
        self.root.title("宽带通查询助手1.0")
        self.root.geometry("860x500")

        self.loadExcel = mtk.LabelFrame(text="导入Excel文件", fg="blue")
        self.loadExcel.place(x=50, y=20, width=220, height=100)
        self.fileChoose = mtk.Button(self.loadExcel, text="选择文件", command=lambda: self.thread_it(self.__loadExcel))
        self.fileChoose.place(x=70, y=5, width=80, height=30)
        self.outExcel = mtk.Button(self.loadExcel, text="导出Excel", command=lambda: self.thread_it(self._saveExcel))
        self.outExcel.place(x=70, y=43, width=80, height=30)

        self.btnBox = mtk.LabelFrame(text="任务栏", fg="blue")
        self.btnBox.place(x=320, y=20, width=220, height=100)
        self.btnStart = mtk.Button(self.btnBox, text="开始", command=lambda: self.thread_it(self.start))
        self.btnStart.place(x=70, y=5, width=80, height=30)
        self.btnEnd = mtk.Button(self.btnBox, text="退出", command=lambda: self.thread_it(self.stop))
        self.btnEnd.place(x=70, y=43, width=80, height=30)

        self.staticData = mtk.LabelFrame(self.root, text="数量统计", fg="blue")
        self.staticData.place(x=583, y=20, width=220, height=100)
        self.labelDataTotal = mtk.Label(self.staticData, text="总导入数量：")
        self.labelDataTotal.place(x=20, y=5, width=80, height=30)
        self.entryDataTotal = mtk.Label(self.staticData, text="0")
        self.entryDataTotal.place(x=120, y=5, width=80, height=30)

        self.labelDoneTotal = mtk.Label(self.staticData, text="已完成数量：")
        self.labelDoneTotal.place(x=20, y=40, width=80, height=30)
        self.entryDoneTotal = mtk.Label(self.staticData,text="0")
        self.entryDoneTotal.place(x=120, y=40, width=80, height=30)

        self.showDataBox = mtk.LabelFrame(text="数据展示区域", fg="blue")
        self.showDataBox.place(x=50, y=150, width=760, height=300)

        title = ['1', '2', '3', '4', '5', '6']
        self.box = ttk.Treeview(self.showDataBox, columns=title, show='headings')
        self.box.place(x=20, y=15, width=730, height=250)
        self.box.column('1', width=80, anchor='center')
        self.box.column('2', width=120, anchor='center')
        self.box.column('3', width=180, anchor='center')
        self.box.column('4', width=80, anchor='center')
        self.box.column('5', width=100, anchor='center')
        self.box.column('6', width=150, anchor='center')
        self.box.heading('1', text='序号')
        self.box.heading('2', text='用户ID')
        self.box.heading('3', text='OLT名称')
        self.box.heading('4', text='PON端口')
        self.box.heading('5', text='用户状态')
        self.box.heading('6', text='时间')
        self.VScroll1 = Scrollbar(self.box, orient='vertical', command=self.box.yview)
        self.VScroll1.pack(side="right", fill="y")
        self.box.configure(yscrollcommand=self.VScroll1.set)

    def deleteTree(self):
        x = self.box.get_children()
        for item in x:
            self.box.delete(item)

    def __loadExcel(self):

        excelPath = filedialog.askopenfilename(title=u'选择文件')
        if excelPath:
            try:
                self.totalData = [["用户ID", "OLT名称", "PON端口", "用户状态", "时间"]]
                self.treeIndex = 1
                self.excelData = []
                self.deleteTree()
                wb = load_workbook(excelPath)
                ws = wb.active
                self.excelData = [i[0] for i in list(ws.values)[1:]]
                self.totals = ws.max_row - 1
                showinfo("提示信息", "共导入{}条数据".format(self.totals))
                self.entryDataTotal.configure(text=f"{self.totals}")
            except:
                showerror("错误信息", "请选择正确格式的Excel文件!")
        else:
            showerror("错误信息", "请导入文件!")

    async def __getContent(self, semaphore, userID):
        link = f"http://220.113.1.44/wap/ShowPic.aspx?userid={userID}"
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Host': '220.113.1.44',
            'Pragma': 'no-cache',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36'
        }
        async with semaphore:
            async with aiohttp.ClientSession(headers=headers) as session:
                async with await session.get(link) as resp:
                    content = await resp.text()
                    await asyncio.sleep(random.uniform(0,2))
                    return content

    async def __crawler(self, semaphore, userID):
        content = await self.__getContent(semaphore, userID)
        print(content)
        print(userID)
        print('-'*100)
        olt = re.findall(r"OLT名称:(.*?)<", content)
        pon = re.findall(r"PON端口:(.*?)<", content)
        status = re.findall(r"用户状态:(.*?)<", content)
        date = re.findall(r"时间:(.*?)<", content)
        treeData = [
            self.treeIndex,
            userID,
            olt[0].replace("&nbsp;", "") if olt else None,
            pon[0].replace("&nbsp;", "") if pon else None,
            status[0].replace("&nbsp;", "") if status else None,
            date[0].replace("&nbsp;", "") if date else None,
        ]
        self.totalData.append(treeData[1:])
        self.box.insert("", "end", values=treeData)
        self.entryDoneTotal.configure(text=self.treeIndex)
        self.treeIndex += 1
        self.box.yview_moveto(1.0)

    def _saveExcel(self):
        if not self.totalData:
            showerror("错误信息", "当前不存在任何数据!")
            return

        excelPath = filedialog.asksaveasfilename(title=u'保存文件', filetypes=[("xlsx", ".xlsx")]) + ".xlsx"
        if excelPath.strip(".xlsx"):
            wb = Workbook()
            ws = wb.active
            for line in self.totalData:
                ws.append(line)
            wb.save(excelPath)
            showinfo("提示信息", "保存成功！")

    async def taskManager(self, dataList, func):
        tasks = []
        sem = asyncio.Semaphore(5)
        for userId in dataList:
            task = asyncio.ensure_future(func(sem, userId))
            tasks.append(task)
        await asyncio.gather(*tasks)

    def start(self):
        if self.totals > 0:
            new_loop = asyncio.new_event_loop()
            asyncio.set_event_loop(new_loop)
            loop = asyncio.get_event_loop()
            loop.run_until_complete(self.taskManager(self.excelData, self.__crawler))
            showinfo("提示信息", "任务结束！")
        else:
            showerror("错误信息", "当前无可查数据!")

    def stop(self):
        os._exit(0)

    @staticmethod
    def thread_it(func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = Application()
    app.run()
