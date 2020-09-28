import random
import threading
import json
import requests
from mttkinter import mtTkinter as mtk
from tkinter import ttk, Scrollbar, filedialog, scrolledtext, END
from tkinter.messagebox import showinfo, showerror
from datetime import datetime
from openpyxl import Workbook, load_workbook
from chaojiying import Chaojiying_Client
import asyncio
import aiohttp
import os
from urllib import parse
from pyquery import PyQuery as pq


class Application(threading.Thread):

    def __init__(self):
        super(Application, self).__init__()
        self.chaojiying = Chaojiying_Client("18612800000", "12345678", "1902")
        self.__createGUI()
        self.lock = threading.Lock()

    def __createGUI(self):
        self.root = mtk.Tk()
        self.root.title("宽带查询助手2.0")
        self.root.geometry("860x700")
        self.companySet = mtk.LabelFrame(text="通信公司选择", fg="blue")
        self.companySet.place(x=50, y=20, width=220, height=100)

        self.labelCompany = mtk.Label(self.companySet, text="通信公司：")
        self.labelCompany.place(x=20, y=20, width=60, height=25)
        self.boxCompany = ttk.Combobox(self.companySet)
        self.boxCompany.place(x=100, y=20, width=80, height=25)
        self.boxCompany["values"] = ["中国移动", "中国联通"]

        self.loadExcel = mtk.LabelFrame(text="导入Excel文件", fg="blue")
        self.loadExcel.place(x=320, y=20, width=220, height=100)
        self.fileChoose = mtk.Button(self.loadExcel, text="选择文件", command=lambda: self.thread_it(self.__loadExcel))
        self.fileChoose.place(x=15, y=40, width=80, height=25)
        self.outExcel = mtk.Button(self.loadExcel, text="导出Excel", command=lambda: self.thread_it(self._saveExcel))
        self.outExcel.place(x=120, y=40, width=80, height=25)

        self.staticData = mtk.Label(self.loadExcel, text="数量统计：")
        self.staticData.place(x=20, y=5, width=60, height=20)
        self.entryShowNum = mtk.Label(self.loadExcel, text="0/0")
        self.entryShowNum.place(x=90, y=5, width=100, height=20)

        self.showDataBox = mtk.LabelFrame(text="数据展示区域", fg="blue")
        self.showDataBox.place(x=50, y=150, width=760, height=300)

        title = ['1', '2', '3', '4', '5', '6', '7', '8']
        self.box = ttk.Treeview(self.showDataBox, columns=title, show='headings')
        self.box.place(x=20, y=15, width=730, height=250)
        self.box.column('1', width=30, anchor='center')
        self.box.column('2', width=90, anchor='center')
        self.box.column('3', width=100, anchor='center')
        self.box.column('4', width=80, anchor='center')
        self.box.column('5', width=60, anchor='center')
        self.box.column('6', width=80, anchor='center')
        self.box.column('7', width=100, anchor='center')
        self.box.column('8', width=80, anchor='center')

        self.box.heading('1', text='序号')
        self.box.heading('2', text='宽带编码')
        self.box.heading('3', text='机主姓名')
        self.box.heading('4', text='产品名称')
        self.box.heading('5', text='签约速率')
        self.box.heading('6', text='联系电话')
        self.box.heading('7', text='装机地址')
        self.box.heading('8', text='到期时间')
        self.VScroll1 = Scrollbar(self.box, orient='vertical', command=self.box.yview)
        self.VScroll1.pack(side="right", fill="y")
        self.box.configure(yscrollcommand=self.VScroll1.set)

        self.btnBox = mtk.LabelFrame(text="任务栏", fg="blue")
        self.btnBox.place(x=590, y=20, width=220, height=100)
        self.btnStart = mtk.Button(self.btnBox, text="开始", command=lambda: self.thread_it(self.start))
        self.btnStart.place(x=20, y=20, width=80, height=30)
        self.btnEnd = mtk.Button(self.btnBox, text="退出", command=lambda: self.thread_it(self.stop))
        self.btnEnd.place(x=120, y=20, width=80, height=30)

        self.loginfo = mtk.LabelFrame(text="log日志信息", fg="blue")
        self.loginfo.place(x=50, y=480, width=760, height=150)

        self.logtext = scrolledtext.ScrolledText(self.loginfo, fg="green")
        self.logtext.place(x=20, y=15, width=730, height=100)

    def deleteTree(self):
        x = self.box.get_children()
        for item in x:
            self.box.delete(item)

    def __loadExcel(self):

        excelPath = filedialog.askopenfilename(title=u'选择文件')
        if excelPath:
            try:
                self.deleteTree()
                self.treeIndex = 1
                self.totals = 0
                self.excelData = []
                self.totalData = [["机主姓名", "宽带编码", "产品名称", "签约速率", "联系电话", "装机地址", "到期时间"]]
                wb = load_workbook(excelPath)
                ws = wb.active
                self.totals = ws.max_row - 1
                self.index = ws.max_column
                if self.index == 2:
                    self.excelData = [[str(i[1]), str(i[0])] for i in list(ws.values)[1:]]
                else:
                    self.excelData = [[str(i[0])] for i in list(ws.values)[1:]]
                showinfo("提示信息", "共导入{}条数据".format(self.totals))
                self.entryShowNum.configure(text=f"{0}/{self.totals}")
                self.logtext.insert(END,
                                    f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t导入Excel文件成功,共计{self.totals}条数据.\n")
                self.logtext.yview_moveto(1.0)
            except Exception as e:
                print(e.args)
                showerror("错误信息", "请选择正确格式的Excel文件!")
        else:
            showerror("错误信息", "请导入文件!")

    async def __getContentFor10010(self, semaphore, userName, userCode):
        async with semaphore:
            link = f"https://openapp.10010.com/bj/kdxf/Kdxf_queryUserProductMes.action?{datetime.now().strftime('%H:%M:%S')}%20GMT+0800%20(%D6%D0%B9%FA%B1%EA%D7%BC%CA%B1%BC%E4)&query_type=1&query_value1=40&query_value2={userCode}&query_value3={userName}"
            headers = {
                'Host': 'openapp.10010.com',
                'Connection': 'keep-alive',
                'Accept': '*/*',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36',
                'X-Requested-With': 'XMLHttpRequest',
                'Sec-Fetch-Site': 'same-origin',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Dest': 'empty',
                'Referer': 'https://openapp.10010.com/bj/kdxf/Kdxf_toKdxf.action',
                'Accept-Encoding': 'gzip, deflate, br',
                'Accept-Language': 'zh-CN,zh;q=0.9'
            }

            async with aiohttp.ClientSession(headers=headers) as session:
                # try:
                async with await session.get(link) as resp:
                    self.logtext.insert(END,
                                        f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userName}\t\t{userCode}\t\t正在查询客户宽带信息,请不要关闭窗口.\n")
                    self.logtext.yview_moveto(1.0)
                    content = await resp.text()
                    print(111111, content)
                    await asyncio.sleep(random.uniform(0, 1))
                    data = json.loads(content.replace("'", "\"")).get("json")
                    return data
                # except Exception as e:
                #     print(22222, e.args)
                #     return

    def __getContentFor10086(self, userCode):
        self.logtext.insert(END,
                            f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userCode}\t\t正在查询客户宽带信息,请不要关闭窗口.\n")
        self.logtext.yview_moveto(1.0)
        link = "https://service.bj.10086.cn/poffice/renewal/mainOfRenewal.action"
        headers = {
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Host': 'service.bj.10086.cn',
            'Referer': 'https://service.bj.10086.cn/poffice/renewal/mainOfRenewal.action',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36',
        }
        session = requests.Session()
        session.headers = headers
        try:
            session.get(link)
            self.logtext.insert(END,
                                f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userCode}\t\t正在进行识别验证码,请不要关闭窗口.\n")
            self.logtext.yview_moveto(1.0)
            content = session.get("https://service.bj.10086.cn/poffice/jsp/package/bb/checkCode.jsp").content
            result = self.chaojiying.PostPic(content, 1902)
            captchaCode = result.get("pic_str")
            pic_id = result.get("pic_id")
            link = f"https://service.bj.10086.cn/poffice/renewal/getBBnumberByMobile.action"
            formData = {
                "BILL_ID": userCode,
                "verificationCode": captchaCode
            }
            captchaRes = session.post(link, data=parse.urlencode(formData))
            if "erroryzm" in captchaRes.text:
                res = self.chaojiying.ReportError(pic_id)
                return
            elif "null" in captchaRes.text:
                return "<html><title>没有查到相关记录</title></html>"
            elif "error" in captchaRes.text:
                self.logtext.insert(END, f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userCode}\t\t发现错误宽带账号.\n")
                return "<html><title>账号输入有误</title></html>"
            else:
                realBill = json.loads(captchaRes.json())[0].get("realBill")
                link = "https://service.bj.10086.cn/poffice/renewal/toRenewalPage.action"
                formData = {
                    "billId": realBill
                }
                self.logtext.insert(END,
                                    f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userCode}\t\t正在解析网页数据,请不要关闭窗口.\n")
                self.logtext.yview_moveto(1.0)
                content = session.post(link, data=parse.urlencode(formData)).text
                return content

        except Exception as e:
            print(e.args)

            return

    def __content10010(self, contentList, userName, userCode):
        treeDataList = []
        if contentList:
            for content in contentList:
                try:
                    treeData = [
                        self.treeIndex,
                        content.get("serialNumber"),
                        content.get("custName"),
                        content.get("productName"),
                        content.get("speed"),
                        content.get("guhuaNumber"),
                        content.get("installAddr"),
                        content.get("endDate").split(" ")[0],
                    ]
                    treeDataList.append(treeData)
                except Exception as e:
                    print(e.args)
                    treeData = [
                        self.treeIndex,
                        userCode,
                        userName,
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                    treeDataList.append(treeData)
                self.logtext.insert(END,
                                    f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userName}\t\t{userCode}\t\t获取数据成功.\n")
                self.logtext.yview_moveto(1.0)

                self.treeIndex += 1

            return treeDataList
        return

    def __content10086(self, content, userCode):
        if content:
            html = pq(content)
            try:
                # "宽带编码", "机主姓名", "产品名称", "签约速率", "联系电话", "装机地址", "到期时间"
                treeData = [
                    self.treeIndex,
                    userCode,
                    html.find("div.currinfos-cont:nth-child(2) dl:nth-child(1) dd").text(),
                    html.find("div.currinfos-cont:nth-child(2) dl:nth-child(4) dd").text(),
                    "",
                    html.find("div.currinfos-cont:nth-child(2) dl:nth-child(3) dd").text(),
                    "",
                    html.find("div.currinfos-cont:nth-child(2) dl:nth-child(5) dd").text(),
                ]
            except:
                treeData = [
                    self.treeIndex,
                    "",
                    userCode,
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            self.logtext.insert(END,
                                f"{datetime.now().strftime('%Y:%m:%d %H:%M:%S')}\t\t{userCode}\t\t获取数据成功.\n")
            self.logtext.yview_moveto(1.0)
            return treeData

        return

    def __crawler10086(self, semaphore, userCode):
        semaphore.acquire()
        _content = self.__getContentFor10086(userCode)
        treeData = self.__content10086(_content, userCode)
        if treeData:
            self.totalData.append(treeData[1:])
            self.box.insert("", "end", values=treeData)
            self.entryShowNum.configure(text=f"{self.treeIndex}/{self.totals}")
            self.lock.acquire()
            self.treeIndex += 1
            self.lock.release()
            self.excelData.remove([userCode])
            self.box.yview_moveto(1.0)
        semaphore.release()

    async def __crawler10010(self, semaphore, userCode, userName):
        _content = await self.__getContentFor10010(semaphore, userName, userCode)
        treeDataList = self.__content10010(_content, userName, userCode)
        if treeDataList:
            for treeData in treeDataList:
                self.totalData.append(treeData[1:])
                self.box.insert("", "end", values=treeData)
                self.entryShowNum.configure(text=f"{treeData[0]}/{self.totals}")
                try:
                    self.excelData.remove([userCode, userName])
                except:
                    pass
                self.box.yview_moveto(1.0)

    def _saveExcel(self):
        if not self.totalData:
            showerror("错误信息", "当前不存在任何数据!")
            return

        excelPath = filedialog.asksaveasfilename(title=u'保存文件', filetypes=[("xlsx", ".xlsx")]) + ".xlsx"
        if excelPath.strip(".xlsx"):
            wb = Workbook()
            ws = wb.active
            for line in self.totalData:
                ws.append(line)
            wb.save(excelPath)
            showinfo("提示信息", "保存成功！")

    async def taskManager10010(self, dataList, func):
        tasks = []
        semaphore = asyncio.Semaphore(5)
        for data in dataList:
            task = asyncio.ensure_future(func(semaphore, *data))
            tasks.append(task)

        await asyncio.wait(tasks)

    def taskManager10086(self, dataList, func):
        semaphore = threading.Semaphore(2)
        ts = [threading.Thread(target=func, args=(semaphore, usrCode[0],)) for usrCode in dataList]
        [t.start() for t in ts]
        [t.join() for t in ts]

    def start(self):
        self.company = self.boxCompany.get()
        if not self.company:
            showerror("错误信息", "请选择通信公司!")
        else:

            if self.totals > 0:
                if self.company == "中国联通":
                    new_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(new_loop)
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(self.taskManager10010(self.excelData, self.__crawler10010))
                else:

                    self.box.column('2', width=100, anchor='center')
                    self.box.column('3', width=60, anchor='center')
                    self.box.column('4', width=170, anchor='center')
                    self.box.column('5', width=0, anchor='center')
                    self.box.column('6', width=120, anchor='center')
                    self.box.column('7', width=0, anchor='center')
                    self.box.column('8', width=160, anchor='center')
                    self.taskManager10086(self.excelData, self.__crawler10086)

            else:
                showerror("错误信息", "当前无用户信息!")

            while self.excelData:
                if self.company == "中国联通":
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(self.taskManager10010(self.excelData, self.__crawler10010))
                else:
                    self.taskManager10086(self.excelData, self.__crawler10086)
            showinfo("提示信息", "任务结束!")

    def stop(self):
        os._exit(0)

    @staticmethod
    def thread_it(func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = Application()
    app.run()
